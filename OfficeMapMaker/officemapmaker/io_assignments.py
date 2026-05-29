"""Load people-to-office assignments from xlsx, legacy xls, or csv.

The loader is **format-agnostic**: a single ``load_assignments(path)`` call
detects the actual container by sniffing the file's first bytes (extensions
lie — the real-world sample is named ``.xlsx`` but is actually legacy BIFF
``.xls``) and returns a normalized ``list[Assignment]``.

The returned records preserve source order and remember each row's
1-based source line number, which Pass 1 (``validate labels``) uses when
reporting errors like ``row 47: office 9999 is not on the map``.

Only the first / active worksheet is read for xlsx and xls.

Header columns required (case-insensitive, whitespace-trimmed):
    - Full Name
    - Office number
    - Team

The header row may appear anywhere within the first 10 rows of the sheet
(real-world spreadsheets often have a title row above the column names).
Empty rows are silently skipped. Cells are ``.strip()``'d. Excel-numeric
office IDs (e.g. ``1480.0``) are normalized to clean strings (``"1480"``);
alphanumeric IDs (e.g. ``"1479A"``) pass through unchanged.
"""

from __future__ import annotations

import csv
from dataclasses import dataclass
from pathlib import Path
from typing import Iterable, Optional


# Header bytes used for format detection.
_XLSX_MAGIC = b"PK\x03\x04"
_XLS_MAGIC = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"

# Logical column names; matched case-insensitively after stripping whitespace.
_COL_NAME = "full name"
_COL_OFFICE = "office number"
_COL_TEAM = "team"
_REQUIRED_COLUMNS = (_COL_NAME, _COL_OFFICE, _COL_TEAM)

# How many rows above the data to search for the header.
_HEADER_SEARCH_DEPTH = 10


@dataclass(frozen=True)
class Assignment:
    """One row of the assignments spreadsheet, normalized.

    Attributes:
        name: Full Name with surrounding whitespace stripped, original case preserved.
        office_id: Office number as a string; numeric Excel values are converted
            ("1480.0" -> "1480"); alphanumeric IDs pass through unchanged.
        team: Team name with surrounding whitespace stripped, original case preserved.
        source_row: 1-based row number in the source file, including the
            header. Useful for error messages.
    """

    name: str
    office_id: str
    team: str
    source_row: int


class AssignmentLoadError(Exception):
    """Raised when the assignments file cannot be loaded or parsed."""


def detect_format(path: Path) -> str:
    """Return ``"xlsx"``, ``"xls"``, or ``"csv"`` based on the file header bytes.

    Extension is **not** consulted — the real-world sample at the heart of
    this project is named ``.xlsx`` but is actually legacy BIFF.
    """
    path = Path(path)
    with path.open("rb") as f:
        head = f.read(8)
    if head[:4] == _XLSX_MAGIC:
        return "xlsx"
    if head[:8] == _XLS_MAGIC:
        return "xls"
    return "csv"


def load_assignments(path: Path | str) -> list[Assignment]:
    """Load assignments from any supported format. See module docstring."""
    path = Path(path)
    if not path.exists():
        raise AssignmentLoadError(f"file not found: {path}")
    if not path.is_file():
        raise AssignmentLoadError(f"not a file: {path}")

    kind = detect_format(path)
    if kind == "xlsx":
        rows = _read_xlsx_rows(path)
    elif kind == "xls":
        rows = _read_xls_rows(path)
    else:
        rows = _read_csv_rows(path)

    return _normalize(rows, source=path)


# ---------------------------------------------------------------------------
# Format-specific readers — each yields (row_index_1_based, [cell, cell, ...])
# ---------------------------------------------------------------------------


def _read_xlsx_rows(path: Path) -> list[tuple[int, list[object]]]:
    try:
        import openpyxl
    except ImportError as e:
        raise AssignmentLoadError(
            "openpyxl is required to read .xlsx files (pip install openpyxl)"
        ) from e

    try:
        wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    except Exception as e:
        raise AssignmentLoadError(f"could not open xlsx {path}: {e}") from e

    try:
        ws = wb.active
        out: list[tuple[int, list[object]]] = []
        for i, row in enumerate(ws.iter_rows(values_only=True), start=1):
            out.append((i, list(row)))
        return out
    finally:
        wb.close()


def _read_xls_rows(path: Path) -> list[tuple[int, list[object]]]:
    try:
        import xlrd
    except ImportError as e:
        raise AssignmentLoadError(
            "xlrd==1.2.0 is required to read legacy .xls files (pip install xlrd==1.2.0)"
        ) from e

    try:
        book = xlrd.open_workbook(str(path))
    except Exception as e:
        # The OLE2 magic also covers rights-managed/encrypted Office files
        # (Azure IRM, password-protected .xlsx, etc.). Those contain an
        # ``EncryptedPackage`` OLE stream that xlrd has no way to decrypt.
        # Detect that case and give a workflow-actionable error.
        if _is_encrypted_office_doc(path):
            raise AssignmentLoadError(
                f"{path} is a password-protected or rights-managed Office document "
                f"(contains an EncryptedPackage stream). Open it in Excel and "
                f"'Save As' an unprotected .xlsx or .csv before running OfficeMapMaker."
            ) from e
        raise AssignmentLoadError(f"could not open xls {path}: {e}") from e

    sheet = book.sheet_by_index(0)
    out: list[tuple[int, list[object]]] = []
    for r in range(sheet.nrows):
        row: list[object] = []
        for c in range(sheet.ncols):
            cell = sheet.cell(r, c)
            # xlrd returns numeric cells as float regardless of integer-ness;
            # _normalize() handles the ".0" cleanup at the office_id step.
            row.append(cell.value)
        out.append((r + 1, row))
    return out


def _is_encrypted_office_doc(path: Path) -> bool:
    """True if the OLE2 file contains an ``EncryptedPackage`` stream.

    Rights-managed Office docs (Azure IRM, ``Encrypt with Password``, etc.) are
    stored as an OLE2 container whose primary payload is an encrypted blob in
    a stream named ``EncryptedPackage``. Stream names appear as UTF-16-LE in
    the OLE2 directory entries.
    """
    try:
        with path.open("rb") as f:
            data = f.read(64 * 1024)
    except OSError:
        return False
    return b"E\x00n\x00c\x00r\x00y\x00p\x00t\x00e\x00d\x00P\x00a\x00c\x00k\x00a\x00g\x00e\x00" in data


def _read_csv_rows(path: Path) -> list[tuple[int, list[object]]]:
    # utf-8-sig transparently strips the BOM that Excel's "Save As CSV" adds.
    try:
        with path.open("r", encoding="utf-8-sig", newline="") as f:
            reader = csv.reader(f)
            return [(i, list(row)) for i, row in enumerate(reader, start=1)]
    except UnicodeDecodeError as e:
        raise AssignmentLoadError(
            f"could not decode csv {path} as utf-8 (try saving as UTF-8): {e}"
        ) from e


# ---------------------------------------------------------------------------
# Normalization
# ---------------------------------------------------------------------------


def _cell_to_str(value: object) -> str:
    """Stringify a cell value, stripping whitespace and Excel's trailing '.0'."""
    if value is None:
        return ""
    if isinstance(value, float):
        # Excel returns ALL numerics as float in xlrd, and often in openpyxl.
        # 1480.0 -> "1480", 12.5 -> "12.5".
        if value.is_integer():
            return str(int(value))
        return str(value).strip()
    if isinstance(value, int):
        return str(value)
    return str(value).strip()


def _find_header(
    rows: list[tuple[int, list[object]]],
) -> Optional[tuple[int, dict[str, int]]]:
    """Locate the header row within the first ``_HEADER_SEARCH_DEPTH`` rows.

    Returns ``(header_row_index, {logical_col_name: cell_index})`` on success,
    or ``None`` if no qualifying header is found.
    """
    for idx, (row_no, cells) in enumerate(rows[:_HEADER_SEARCH_DEPTH]):
        col_map: dict[str, int] = {}
        for ci, cell in enumerate(cells):
            label = _cell_to_str(cell).lower().strip()
            if label in _REQUIRED_COLUMNS and label not in col_map:
                col_map[label] = ci
        if all(c in col_map for c in _REQUIRED_COLUMNS):
            return row_no, col_map
    return None


def _normalize(
    rows: list[tuple[int, list[object]]], source: Path
) -> list[Assignment]:
    if not rows:
        raise AssignmentLoadError(f"no rows found in {source}")

    found = _find_header(rows)
    if found is None:
        raise AssignmentLoadError(
            f"could not find a header row containing all of "
            f"{_REQUIRED_COLUMNS!r} within the first {_HEADER_SEARCH_DEPTH} rows of {source}"
        )
    header_row_no, col_map = found

    out: list[Assignment] = []
    for row_no, cells in rows:
        if row_no <= header_row_no:
            continue

        name = _cell_to_str(_safe_get(cells, col_map[_COL_NAME]))
        office = _cell_to_str(_safe_get(cells, col_map[_COL_OFFICE]))
        team = _cell_to_str(_safe_get(cells, col_map[_COL_TEAM]))

        if not (name or office or team):
            # Fully empty row — silently skip.
            continue

        out.append(
            Assignment(name=name, office_id=office, team=team, source_row=row_no)
        )

    if not out:
        raise AssignmentLoadError(f"no data rows found in {source} after the header")

    return out


def _safe_get(cells: list[object], idx: int) -> object:
    return cells[idx] if idx < len(cells) else ""


# Public re-exports.
__all__ = [
    "Assignment",
    "AssignmentLoadError",
    "detect_format",
    "load_assignments",
]
